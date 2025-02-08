'''Напишіть скрапер, який отримає список всіх ноутбуків із сайту,
 включаючи їх детальну інформацію. Для роботи використовуємо бібліотеки bs4, requests.'''

import requests 
import csv
from bs4 import BeautifulSoup as bs
import openpyxl as op


from Config import url_tamplat,teg_name,teg_description,teg_div,class_p,col

def parser(url_tamplate,teg_div,class_p,teg_name,teg_description):
    
    
    r=requests.get(url_tamplate)#Отправляем запрос на получение страницы
    
    soup = bs(r.text, "lxml")#Собираем все теги со страницы 
    
    produkts = soup.find_all(teg_div, class_= class_p)#Поиск нужного блока по и их запись 
   
    #Сщздаем пустой словарь и словник для записи данных
    laptops=[]
    laptop={}
  #Перебираем блоки для выбора нужных даных
    for produkt in produkts:
        #Отбираем название и записываем в словарь
        laptop['name']=produkt.find(teg_name).text  
        #Отбирпем описание и записываем в словарь
        laptop['description']=produkt.find(class_= teg_description).text  
        laptops.append(laptop) # Записываем словари в список
        
    # Записываем в нужные форматы
    writer_csv('Laptops_csv.csv',laptops)
    writer_xlsx('Laptops_XLSX.xlsx',laptops)

def writer_csv(filename,data):
    #Заголовки 
    headers = ['name', 'description']
    
# Открываем файл на запись
    with open(filename, 'a+', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, headers)

        # Проверяем, пустой ли файл
        csvfile.seek(0)
        reader = csv.reader(csvfile)
        try:
            first_row = next(reader)
            if not first_row:  # Файл пустой или только что создан
                writer.writeheader()
        except StopIteration:  # Файл пустой или только что создан
            writer.writeheader()

        # Перемещаемся в низ файла
        csvfile.seek(0, 2)

        # Записываем данные
        writer.writerows(data)

def writer_xlsx(filename, data):
    # Создаем новый рабочий лист
    workbook = None
    # проверяем на наличие файла для записи
    try:#если файл существует
        workbook = op.load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:# создаем если нет
        workbook = op.Workbook()
        sheet = workbook.active

    # Определяем начальную строку для записи
   
    headers = ['name', 'description']
    if not headers:
        headers = data[0].keys()

    # Записываем заголовки в первую строку
    for col, col_name in enumerate(headers):
        sheet.cell(row=1, column=col+1, value=col_name)
    start_row = sheet.max_row + 1
    # Записываем полученые  данные
    for row_index, row_data in enumerate(data, start=start_row):
        for col_index, value in enumerate(row_data.values()):
            sheet.cell(row=row_index, column=col_index+1, value=value)

    workbook.save(filename)

def scrol(url_tamplate,teg_name,teg_description,teg_div,class_p,col):
    for i in range(1,col+1):# Перебираем URL для парсинга
        url=url_tamplate+str(i)
        parser(url,teg_name,teg_description,teg_div,class_p)# Парсим полученый URL

        
if __name__== "__main__":
    scrol(url_tamplat,teg_div,class_p,teg_name,teg_description,col)