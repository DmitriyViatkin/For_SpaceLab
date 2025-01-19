import requests 
from bs4 import BeautifulSoup as bs
import openpyxl as op


from Config import url_tamplat,FILE_NAME,sheet_name




def find_last_non_empty_row(file_path, sheet_name=None):
    # Открываем Excel-файл
    wb = op.load_workbook(file_path)
    sheet = wb[sheet_name] if sheet_name else wb.active

    # Начинаем с последней строки
    last_row = sheet.max_row
    while last_row > 0:
        # Проверяем, есть ли данные в строке
        if any(sheet.cell(row=last_row, column=col).value for col in range(1, sheet.max_column + 1)):
            break
        last_row -= 1  # Переходим на строку выше

    wb.close()
    return last_row
def parser(url_tamplate,FILE_NAME,sheet_name):
    
    print(url_tamplate)
    r=requests.get(url_tamplate)
    
    soup = bs(r.text, "lxml")
    #print(soup)
    produkts = soup.find_all('div', class_= 'product-wrapper card-body')
    i=1
    wb=op.load_workbook(FILE_NAME)
    ws=wb.active
    c=find_last_non_empty_row(FILE_NAME, sheet_name)
    i=0
    for produkt in produkts:
        i+=1
        
        value1=produkt.find('a').text
        ws['A'+str(c+i)]=value1
        
        
        value2 = produkt.find(class_='description card-text').text
        ws['B'+str(c+i)]=value2
        
        
    wb.save('Example.xlsx')

def scrol(url_tamplate):
    for i in range(1,21):
        url=url_tamplate+str(i)
        parser(url,FILE_NAME,sheet_name)

scrol(url_tamplat)